"""
Main validation orchestrator using CrewAI
"""
from typing import Dict, List, Any
import pandas as pd
from crewai import Crew, Task, Process
from langchain.llms import Ollama

from .config import OLLAMA_BASE_URL, OLLAMA_MODEL, OLLAMA_TIMEOUT, OLLAMA_KEEP_ALIVE, OLLAMA_TEMPERATURE, COLUMN_NAMES, INAT_MCP_URL
from .agents import (
    ZoneValidator, CountryValidator, StateValidator, CountyValidator,
    FamilyValidator, GenusValidator, SpeciesValidator, SubspeciesValidator,
    FirstDateValidator, LastDateValidator, YearValidator,
    StateRecordValidator, CountyRecordValidator,
    LocationValidator, NameValidator, CommentValidator,
    RecordQAAgent
)
from .integrations import INatValidator


class LepSocValidationCrew:
    """Main CrewAI orchestrator for validation"""

    def __init__(self, ollama_url: str = OLLAMA_BASE_URL, ollama_model: str = OLLAMA_MODEL,
                 inat_url: str = INAT_MCP_URL, use_inat: bool = True):
        # Initialize Ollama LLM with timeout, keep_alive, and temperature settings
        self.llm = Ollama(
            model=ollama_model,
            base_url=ollama_url,
            timeout=OLLAMA_TIMEOUT,  # Timeout in seconds
            keep_alive=OLLAMA_KEEP_ALIVE,  # Keep model loaded in memory (e.g., "10m")
            temperature=OLLAMA_TEMPERATURE  # Lower temperature reduces hallucinations
        )

        # Pre-load model into memory with a warm-up call
        try:
            print(f"Pre-loading model {ollama_model}...")
            self.llm.invoke("warmup")  # Simple call to load model into memory
            print("✓ Model pre-loaded and ready")
        except Exception as e:
            print(f"⚠ Warning: Could not pre-load model: {e}")

        # Initialize iNaturalist validator (shared across all validators)
        # Use mock_mode if iNat is disabled or unavailable
        self.inat_validator = INatValidator(server_url=inat_url, mock_mode=not use_inat)

        # Create all validation agents
        self.validators = self._create_validators()

        # Initialize QA agent for final cross-row validations
        self.qa_agent = RecordQAAgent()

        # Column mapping
        self.column_names = COLUMN_NAMES

    def _create_validators(self) -> List:
        """Create all 16 validation agents

        - Deterministic validators (7): Zone, Country, State, FirstDate, LastDate, Name, Year
        - iNat-powered validators (7): Family, Genus, Species, Subspecies, County, StateRecord, CountyRecord
        - LLM-powered validators (2): Location shortening, Comment standardization
        """
        validators = [
            # Geographic (deterministic)
            ZoneValidator(),
            CountryValidator(),
            StateValidator(),

            # Taxonomic (AI-powered for iNat integration)
            FamilyValidator(self.llm, self.inat_validator),
            GenusValidator(self.llm, self.inat_validator),
            SpeciesValidator(self.llm, self.inat_validator),
            SubspeciesValidator(self.llm, self.inat_validator),

            # Geographic (deterministic, but with iNat location validation)
            CountyValidator(self.inat_validator),

            # Records (iNat-powered for record status verification)
            StateRecordValidator(self.inat_validator),
            CountyRecordValidator(self.inat_validator),

            # Metadata (location uses LLM for auto-shortening)
            LocationValidator(self.llm),

            # Temporal (deterministic)
            FirstDateValidator(),
            LastDateValidator(),

            # Metadata (name is deterministic)
            NameValidator(),

            # Comments (AI-powered for shortening/standardization)
            CommentValidator(self.llm),

            # Temporal (deterministic)
            YearValidator()
        ]
        return validators

    def validate_row(self, row_index: int, row_data: pd.Series) -> Dict[str, Any]:
        """
        Validate a single row using all agents

        Args:
            row_index: Row index in the dataframe
            row_data: Pandas Series containing row data

        Returns:
            Dict containing validation results
        """
        # Convert row to dict for easier access
        row_dict = row_data.to_dict()

        # Results container
        validation_results = {
            'row_index': row_index,
            'is_valid': True,
            'errors': [],
            'warnings': [],
            'corrections': {},
            'metadata': {},
            'needs_review': False,
            'field_results': {}  # Store individual field results for coloring
        }

        # Run validators directly (no CrewAI Crew needed - validators are simple Python classes)
        try:
            # Run validators directly
            for i, (col_name, validator) in enumerate(zip(self.column_names, self.validators)):
                value = row_data.iloc[i] if i < len(row_data) else None

                # Run validation
                result = validator.validate(value, row_dict)

                # Store field result for coloring logic
                validation_results['field_results'][col_name] = result

                # Process results
                if not result.is_valid:
                    validation_results['is_valid'] = False
                    validation_results['errors'].extend(
                        [f"{col_name}: {e}" for e in result.errors]
                    )

                if result.warnings:
                    validation_results['warnings'].extend(
                        [f"{col_name}: {w}" for w in result.warnings]
                    )

                if result.correction is not None:
                    validation_results['corrections'][col_name] = result.correction

                if result.metadata:
                    validation_results['metadata'][col_name] = result.metadata

                # Check if review needed
                if result.metadata.get('needs_inat_check') or \
                   result.metadata.get('needs_inat_verification'):
                    validation_results['needs_review'] = True

        except Exception as e:
            validation_results['is_valid'] = False
            validation_results['errors'].append(f"Validation error: {str(e)}")

        return validation_results

    def validate_file(self, filepath: str, output_path: str = None) -> pd.DataFrame:
        """
        Validate entire Excel/CSV file

        Args:
            filepath: Path to input file
            output_path: Optional path for output file

        Returns:
            DataFrame with validation results
        """
        # Read file
        if filepath.endswith('.xlsx'):
            df = pd.read_excel(filepath, header=None)
        else:
            df = pd.read_csv(filepath, header=None)

        # Detect and skip header row (if first row contains column names)
        if df.iloc[0].astype(str).str.contains('Zone|Country|State|Family|Genus', case=False, na=False).any():
            print("Detected header row, skipping...")
            df = df.iloc[1:].reset_index(drop=True)

        # Handle fewer than 16 columns by padding with empty values
        if len(df.columns) < 16:
            print(f"File has {len(df.columns)} columns, padding to 16...")
            for i in range(len(df.columns), 16):
                df[i] = ''

        # Rename columns
        df.columns = self.column_names

        # Validation results
        all_results = []
        valid_indices = []  # Track which rows are not blank

        print(f"Validating {len(df)} rows...")

        # Process each row
        for index, row in df.iterrows():
            # Skip blank rows (all key fields are empty/NaN)
            key_fields = ['Family', 'Genus', 'Species']
            if all(pd.isna(row.get(field)) or str(row.get(field, '')).strip() == '' for field in key_fields):
                print(f"\nValidating row {index + 1}/{len(df)} - Skipping blank row")
                continue

            valid_indices.append(index)  # Track non-blank rows
            print(f"\nValidating row {index + 1}/{len(df)}")
            result = self.validate_row(index, row)
            all_results.append(result)

            # Show results
            if result['errors']:
                print(f"  ✗ Errors: {', '.join(result['errors'][:3])}")
            if result['warnings']:
                print(f"  ⚠ Warnings: {', '.join(result['warnings'][:3])}")
            if result['corrections']:
                print(f"  ✓ Corrections: {len(result['corrections'])} fields")

        # Filter dataframe to only include non-blank rows
        df_filtered = df.loc[valid_indices].reset_index(drop=True)

        # Run final QA checks for cross-row validation rules
        print("\n" + "="*50)
        print("Running final QA checks...")
        print("="*50)
        all_results = self.qa_agent.validate_record_uniqueness(df_filtered, all_results)
        print("✓ QA checks complete")

        # Apply corrections and create dual-column output
        validated_df = self._apply_corrections(df_filtered, all_results)

        # Save output with color coding
        if output_path:
            if output_path.endswith('.xlsx'):
                self._save_excel_with_colors(df_filtered, validated_df, all_results, output_path)
            else:
                validated_df.to_csv(output_path, index=False)
            print(f"\n✓ Validated file saved to: {output_path}")

        # Summary
        self._print_summary(all_results)

        return validated_df

    def _apply_corrections(self, df: pd.DataFrame, results: List[Dict]) -> pd.DataFrame:
        """Apply corrections to create corrected dataframe"""
        validated_df = df.copy()

        for i, result in enumerate(results):
            # Apply corrections
            for field, correction in result['corrections'].items():
                if field in validated_df.columns:
                    validated_df.loc[i, field] = correction

        return validated_df

    def _save_excel_with_colors(self, original_df: pd.DataFrame, corrected_df: pd.DataFrame,
                                 results: List[Dict], output_path: str):
        """Save Excel file with corrected columns (left) and original columns (right) with color coding

        Color coding (left side only):
        - Red: Error with no correction available
        - Orange: Correction applied
        - Yellow: Warning only
        """
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill

        wb = Workbook()
        ws = wb.active

        # Define colors
        RED_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # Light red
        ORANGE_FILL = PatternFill(start_color="FFD699", end_color="FFD699", fill_type="solid")  # Light orange
        YELLOW_FILL = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # Light yellow

        # Write headers (duplicated: corrected + original)
        headers = list(corrected_df.columns)
        ws.append(headers + headers)  # Double the headers

        # Write data rows
        for idx in range(len(corrected_df)):
            # Get row data
            corrected_row = corrected_df.iloc[idx].tolist()
            original_row = original_df.iloc[idx].tolist()

            # Append combined row (corrected + original)
            ws.append(corrected_row + original_row)

            # Apply color coding to corrected columns (left side only)
            result = results[idx]
            row_num = idx + 2  # +1 for header, +1 for 1-based indexing

            for col_idx, field in enumerate(headers):
                cell = ws.cell(row=row_num, column=col_idx + 1)

                # Determine color based on validation result
                color = None

                # Check if field has validation results
                if result['field_results'].get(field):
                    field_result = result['field_results'][field]

                    # Precedence: Red > Orange > Yellow
                    # Red: Error with no correction
                    if not field_result.is_valid and field not in result['corrections']:
                        color = RED_FILL
                    # Orange: Actual correction applied (not just normalization)
                    elif field in result['corrections'] and field_result.correction_type == "correction":
                        color = ORANGE_FILL
                    # Yellow: Warning only (valid but has warnings)
                    elif field_result.is_valid and field_result.warnings:
                        color = YELLOW_FILL

                # Apply color
                if color:
                    cell.fill = color

        # Save workbook
        wb.save(output_path)

    def _print_summary(self, results: List[Dict]):
        """Print validation summary"""
        total = len(results)
        passed = sum(1 for r in results if r['is_valid'] and not r['corrections'])
        corrected = sum(1 for r in results if r['corrections'])
        failed = sum(1 for r in results if not r['is_valid'])

        # Needs review = rows with errors, warnings, OR real corrections (not just normalizations)
        needs_review = sum(1 for r in results if
            not r['is_valid'] or  # Has errors (red cells)
            r['errors'] or        # Has errors
            r['warnings'] or      # Has warnings (yellow cells)
            any(r['field_results'].get(field, {}).correction_type == 'correction'
                for field in r.get('corrections', {}))  # Has real corrections (orange cells), not just normalizations
        )

        print("\n" + "="*50)
        print("VALIDATION SUMMARY")
        print("="*50)
        print(f"Total Rows: {total}")
        print(f"✓ Passed: {passed} ({passed/total*100:.1f}%)")
        print(f"✓ Corrected: {corrected} ({corrected/total*100:.1f}%)")
        print(f"✗ Failed: {failed} ({failed/total*100:.1f}%)")
        print(f"? Needs Review: {needs_review} ({needs_review/total*100:.1f}%)")
        print("="*50)
